#!/usr/bin/env python3
"""巡检表自动化：改日期 + 发邮件"""
import openpyxl
import smtplib
import sys
import os
import re
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# ====== SMTP 配置 ======
SMTP_SERVER = "smtp.mxhichina.com"
SMTP_PORT = 465
SMTP_USER = "renzhicai@lnitv.com"
SMTP_PASS = "CcZvi1xqXmfYlZ0b"

# ====== 收件人列表 ======
RECIPIENTS = [
    "lifangfang@lnitv.com",
    "renzhicai@lnitv.com",
    "wujian@lnitv.com",
    "xukaixuan@lnitv.com",
    "gezhijia@lnitv.com",
    "liuweiwang@lnitv.com",
]


def modify_dates(filepath, new_date, shift):
    """修改 Excel 里的日期，返回新文件路径"""
    wb = openpyxl.load_workbook(filepath)

    # new_date: "20260610", shift: "白" or "夜"
    today_str = new_date  # 20260610
    month_day = f"{today_str[4:6]}月{today_str[6:8]}日"  # 06月10日
    month_day_short = f"{int(today_str[4:6])}月{int(today_str[6:8])}日"  # 6月10日

    # 新日期各种格式
    today_slash = f"2026/{int(today_str[4:6])}/{int(today_str[6:8])}"  # 2026/6/10

    # 替换模式
    replacements = [
        # 中文日期
        (r'\d{1,2}月\d{1,2}日', month_day_short),
        (r'\d{2}月\d{2}日', month_day),
        # 数字日期
        (r'2026\d{4}', today_str),
        # 斜杠格式日期: 2026/6/2 → 2026/6/10
        (r'2026/\d{1,2}/\d{1,2}', today_slash),
        # 日期+班次（中文）
        (r'\d{1,2}月\d{1,2}白', f'{month_day_short}白'),
        (r'\d{1,2}月\d{1,2}夜', f'{month_day_short}夜'),
        (r'\d{2}月\d{2}白', f'{month_day}白'),
        (r'\d{2}月\d{2}夜', f'{month_day}夜'),
        # 日期+班次（斜杠）: 2026/6/2白 → 2026/6/10白
        (r'2026/\d{1,2}/\d{1,2}白', f'{today_slash}白'),
        (r'2026/\d{1,2}/\d{1,2}夜', f'{today_slash}夜'),
    ]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                # 处理字符串
                if isinstance(cell.value, str):
                    for pattern, replacement in replacements:
                        cell.value = re.sub(pattern, replacement, cell.value)
                    # 白班：IPTV频道数 141 → 142（文本类型）
                    if shift == "白":
                        cell.value = re.sub(r'iptv[：:]\s*141', 'iptv：142', cell.value, flags=re.IGNORECASE)
                        if cell.value.strip() == '141':
                            cell.value = '142'
                # 白班：IPTV频道数 141 → 142（数字类型）
                if shift == "白" and isinstance(cell.value, (int, float)) and int(cell.value) == 141:
                    cell.value = 142

    # 新文件名
    dirname = os.path.dirname(filepath)
    basename = os.path.basename(filepath)
    new_basename = re.sub(r'2026\d{4}', today_str, basename)
    new_path = os.path.join(dirname, new_basename)

    wb.save(new_path)
    return new_path


def send_mail(to_list, subject, body, attachments):
    """发送邮件"""
    import email.utils
    msg = MIMEMultipart()
    msg["From"] = SMTP_USER
    msg["To"] = ", ".join(to_list)
    msg["Subject"] = subject
    msg["Date"] = email.utils.formatdate(localtime=True)
    msg["Message-ID"] = email.utils.make_msgid()

    # 正文不能为空，否则容易进垃圾箱
    if not body or not body.strip():
        body = "请查收附件。"
    msg.attach(MIMEText(body + "\n\n--\n任志财", "plain", "utf-8"))

    for filepath in attachments:
        filename = os.path.basename(filepath)
        with open(filepath, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", 'attachment', filename=filename)
            msg.attach(part)

    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
        server.login(SMTP_USER, SMTP_PASS)
        server.sendmail(SMTP_USER, to_list, msg.as_string())

    print(f"✅ 已发送到: {', '.join(to_list)}")


import subprocess as sp

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("用法: python3 auto-inspection.py <文件路径> <日期YYYYMMDD> [白|夜] [--now]")
        print("      --now  立即发送（不加则定时发送：白班当天21:00，夜班次日9:00）")
        sys.exit(1)

    # --send-only 模式：只发文件
    if sys.argv[1] == "--send-only":
        f = sys.argv[2]
        send_mail(RECIPIENTS, os.path.basename(f), "", [f])
        print(f"✅ 已发送: {os.path.basename(f)}")
        sys.exit(0)

    filepath = sys.argv[1]
    new_date = sys.argv[2]
    shift = sys.argv[3] if len(sys.argv) > 3 and sys.argv[3] != "--now" else "白"
    send_now = "--now" in sys.argv

    print(f"📝 修改日期为 {new_date} ({shift}班)")
    new_file = modify_dates(filepath, new_date, shift)
    filename = os.path.basename(new_file)
    print(f"📄 生成文件: {new_file}")

    # 计算发送时间
    import datetime
    y, m, d = int(new_date[:4]), int(new_date[4:6]), int(new_date[6:8])
    if shift == "白":
        send_time = datetime.datetime(y, m, d, 21, 0)
    else:
        next_day = datetime.datetime(y, m, d) + datetime.timedelta(days=1)
        send_time = next_day.replace(hour=9, minute=0)

    if send_now:
        print(f"📧 立即发送邮件...")
        send_mail(RECIPIENTS, filename, "", [new_file])
        print("🎉 全部完成!")
    else:
        at_time = send_time.strftime("%H:%M %Y-%m-%d")
        script = os.path.abspath(__file__)
        abs_path = os.path.abspath(new_file)
        cmd = f'/usr/bin/python3 {script} --send-only "{abs_path}"'
        sp.run(["at", at_time], input=cmd.encode(), capture_output=True, timeout=10)
        print(f"⏰ 已设置定时发送: {send_time.strftime('%Y-%m-%d %H:%M')} ({shift}班)")
        print(f"   到点自动发到: {', '.join(RECIPIENTS)}")
